# src/ingestion/processors/excel_processor.py
"""
Enhanced Excel Processor with Azure AI Integration
Handles complex Excel files including embedded Visio diagrams, charts, and images
"""
import os
import logging
import hashlib
import json
import tempfile
import base64
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple, Union
from datetime import datetime
import zipfile
from io import BytesIO
import numpy as np

# Excel processing libraries
try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Image processing
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Base processor import
try:
    from .base_processor import BaseProcessor
except ImportError:
    # Fallback if base processor doesn't exist yet
    class BaseProcessor:
        def __init__(self, config=None):
            self.config = config or {}
            self.logger = logging.getLogger(__name__)


class NumpyEncoder(json.JSONEncoder):
    """Custom JSON encoder for NumPy types"""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)


class ExcelProcessor(BaseProcessor):
    """Process Excel files with embedded objects and Azure AI integration"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None, azure_client=None):
        """Initialize Excel processor"""
        super().__init__(config)
        self.azure_client = azure_client
        
        # Supported formats
        self.supported_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']
        
        # Processing settings
        self.extract_images = self.config.get('extract_images', True)
        self.extract_charts = self.config.get('extract_charts', True)
        self.extract_embedded_objects = self.config.get('extract_embedded_objects', True)
        self.process_formulas = self.config.get('process_formulas', True)
        self.max_sheet_size = self.config.get('max_sheet_size', 1000000)  # cells
        
        # Floor manager data patterns
        self.floor_patterns = ['floor', 'level', 'story', 'storey']
        self.manager_patterns = ['manager', 'supervisor', 'lead', 'head']
        
        self.logger.info("Excel processor initialized with Azure AI support")
    
    def can_process(self, file_path: str) -> bool:
        """Check if file can be processed by this processor"""
        return Path(file_path).suffix.lower() in self.supported_extensions
    
    def process(self, file_path: str, metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Process Excel file and extract all content"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing")
        
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        self.logger.info(f"📊 Starting Excel processing: {file_path}")
        self.logger.info(f"📊 File type: EXCEL | File size: {file_path.stat().st_size:,} bytes")
        self.logger.info(f"📊 Processor: ExcelProcessor")
        
        try:
            # Initialize result structure
            result = {
                'status': 'success',
                'file_path': str(file_path),
                'file_name': file_path.name,
                'sheets': [],
                'embedded_objects': [],
                'charts': [],
                'images': [],
                'hierarchical_data': {},
                'metadata': {
                    'processor': 'excel',
                    'timestamp': datetime.now().isoformat(),
                    'file_size': file_path.stat().st_size,
                    **(metadata or {})
                },
                'chunks': []
            }
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
            
            # Extract workbook properties
            result['metadata']['properties'] = self._extract_workbook_properties(workbook)
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet_data = self._process_sheet(workbook[sheet_name])
                result['sheets'].append(sheet_data)
            
            # Extract embedded objects
            if self.extract_embedded_objects:
                embedded_objects = self._extract_embedded_objects(file_path, workbook)
                result['embedded_objects'] = embedded_objects
            
            # Extract charts
            if self.extract_charts:
                charts = self._extract_charts(workbook)
                result['charts'] = charts
            
            # Extract images
            if self.extract_images:
                images = self._extract_images(workbook)
                result['images'] = images
            
            # Extract hierarchical/floor manager data
            hierarchical_data = self._extract_hierarchical_data(workbook)
            if hierarchical_data:
                result['hierarchical_data'] = hierarchical_data
            
            # Analyze sheet relationships
            relationships = self._extract_sheet_relationships(workbook)
            result['metadata']['sheet_relationships'] = relationships
            
            # Create content chunks for vector storage
            chunks = self._create_chunks(result)
            result['chunks'] = chunks
            
            self.logger.info(f"Successfully processed Excel file with {len(result['sheets'])} sheets, "
                           f"{len(result['embedded_objects'])} embedded objects, "
                           f"{len(result['images'])} images")
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error processing Excel file: {e}")
            return {
                'status': 'error',
                'error': str(e),
                'file_path': str(file_path)
            }
    
    def _extract_workbook_properties(self, workbook) -> Dict[str, Any]:
        """Extract workbook properties and metadata"""
        properties = {}
        
        try:
            if hasattr(workbook, 'properties'):
                props = workbook.properties
                properties = {
                    'title': getattr(props, 'title', None),
                    'creator': getattr(props, 'creator', None),
                    'created': str(getattr(props, 'created', None)),
                    'modified': str(getattr(props, 'modified', None)),
                    'last_modified_by': getattr(props, 'lastModifiedBy', None),
                    'description': getattr(props, 'description', None),
                    'keywords': getattr(props, 'keywords', None),
                    'category': getattr(props, 'category', None),
                    'company': getattr(props, 'company', None)
                }
        except Exception as e:
            self.logger.warning(f"Could not extract workbook properties: {e}")
        
        return properties
    
    def _process_sheet(self, sheet) -> Dict[str, Any]:
        """Process a single Excel sheet"""
        sheet_data = {
            'name': sheet.title,
            'dimensions': f"{sheet.min_row}:{sheet.max_row}x{sheet.min_column}:{sheet.max_column}",
            'data': [],
            'formulas': [],
            'merged_cells': [],
            'comments': [],
            'hyperlinks': []
        }
        
        # Check sheet size
        total_cells = (sheet.max_row - sheet.min_row + 1) * (sheet.max_column - sheet.min_column + 1)
        if total_cells > self.max_sheet_size:
            self.logger.warning(f"Sheet {sheet.title} too large ({total_cells} cells), sampling data")
            sample_rows = min(sheet.max_row, 1000)
        else:
            sample_rows = sheet.max_row
        
        # Extract cell data
        for row in sheet.iter_rows(min_row=1, max_row=sample_rows):
            row_data = []
            for cell in row:
                cell_info = {
                    'value': cell.value,
                    'coordinate': cell.coordinate
                }
                
                # Extract formula if present
                if self.process_formulas and hasattr(cell, 'formula') and cell.formula:
                    sheet_data['formulas'].append({
                        'cell': cell.coordinate,
                        'formula': cell.formula
                    })
                
                # Extract comment if present
                if cell.comment:
                    sheet_data['comments'].append({
                        'cell': cell.coordinate,
                        'author': cell.comment.author,
                        'text': cell.comment.text
                    })
                
                # Extract hyperlink if present
                if cell.hyperlink:
                    sheet_data['hyperlinks'].append({
                        'cell': cell.coordinate,
                        'target': cell.hyperlink.target,
                        'display': cell.hyperlink.display
                    })
                
                row_data.append(cell_info['value'])
            
            sheet_data['data'].append(row_data)
        
        # Extract merged cells
        for merged_range in sheet.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Convert to DataFrame for easier analysis if pandas available
        if PANDAS_AVAILABLE and sheet_data['data']:
            try:
                df = pd.DataFrame(sheet_data['data'])
                # Convert NumPy types to Python native types
                sheet_data['summary'] = {
                    'rows': int(len(df)),
                    'columns': int(len(df.columns)),
                    'non_null_cells': int(df.count().sum()),
                    'data_types': {str(k): int(v) for k, v in df.dtypes.value_counts().to_dict().items()}
                }
            except Exception as e:
                self.logger.warning(f"Could not create DataFrame summary: {e}")
        
        return sheet_data
    
    def _extract_embedded_objects(self, file_path: Path, workbook) -> List[Dict[str, Any]]:
        """Extract embedded objects from Excel file"""
        embedded_objects = []
        
        try:
            # Excel files are actually zip archives
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                # Look for embedded objects in the Excel structure
                embeddings_path = 'xl/embeddings/'
                
                for file_info in zip_file.filelist:
                    if file_info.filename.startswith(embeddings_path):
                        try:
                            # Extract embedded file
                            embedded_data = zip_file.read(file_info.filename)
                            
                            # Determine object type
                            obj_type = self._identify_embedded_type(file_info.filename, embedded_data)
                            
                            obj_info = {
                                'filename': file_info.filename,
                                'type': obj_type,
                                'size': len(embedded_data)
                            }
                            
                            # Process based on type
                            if obj_type == 'visio' and self.azure_client:
                                # Process Visio diagram with Azure AI
                                ocr_result = self.azure_client.process_image(embedded_data, image_type='diagram')
                                if ocr_result['success']:
                                    obj_info['extracted_text'] = ocr_result['text']
                                    obj_info['ocr_regions'] = ocr_result['regions']
                                else:
                                    obj_info['extraction_error'] = ocr_result['error']
                            
                            elif obj_type in ['image', 'diagram'] and self.azure_client:
                                # Process other images
                                ocr_result = self.azure_client.process_image(embedded_data)
                                if ocr_result['success']:
                                    obj_info['extracted_text'] = ocr_result['text']
                                
                                # Also analyze image content
                                analysis = self.azure_client.analyze_image(embedded_data)
                                if analysis['success']:
                                    obj_info['image_analysis'] = analysis['analysis']
                            
                            # Store raw data for potential future processing
                            obj_info['data'] = base64.b64encode(embedded_data).decode('utf-8')
                            
                            embedded_objects.append(obj_info)
                            
                        except Exception as e:
                            self.logger.error(f"Error extracting embedded object {file_info.filename}: {e}")
                
                # Also check for OLE objects in worksheets
                for sheet in workbook.worksheets:
                    if hasattr(sheet, '_rels') and sheet._rels:
                        for rel in sheet._rels.values():
                            if 'oleObject' in rel.Type:
                                embedded_objects.append({
                                    'sheet': sheet.title,
                                    'type': 'ole_object',
                                    'relationship_type': rel.Type,
                                    'target': rel.Target
                                })
        
        except Exception as e:
            self.logger.error(f"Error extracting embedded objects: {e}")
        
        return embedded_objects
    
    def _identify_embedded_type(self, filename: str, data: bytes) -> str:
        """Identify the type of embedded object"""
        # Check file signature
        if data.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):  # OLE2 signature
            # Could be Visio, check further
            if b'Visio' in data[:1000] or b'Microsoft Visio' in data[:1000]:
                return 'visio'
            return 'ole_object'
        
        # Check for common image formats
        if data.startswith(b'\xff\xd8\xff'):  # JPEG
            return 'image'
        elif data.startswith(b'\x89PNG'):  # PNG
            return 'image'
        elif data.startswith(b'GIF87a') or data.startswith(b'GIF89a'):  # GIF
            return 'image'
        elif data.startswith(b'BM'):  # BMP
            return 'image'
        
        # Check for PDF
        if data.startswith(b'%PDF'):
            return 'pdf'
        
        # Default to binary object
        return 'binary_object'
    
    def _extract_charts(self, workbook) -> List[Dict[str, Any]]:
        """Extract and describe charts from workbook"""
        charts = []
        
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_charts'):
                for idx, chart in enumerate(sheet._charts):
                    chart_info = {
                        'sheet': sheet.title,
                        'index': idx,
                        'type': type(chart).__name__
                    }
                    
                    # Extract chart properties
                    if hasattr(chart, 'title') and chart.title:
                        chart_info['title'] = str(chart.title)
                    
                    # Generate chart description
                    description = self._describe_chart(chart)
                    chart_info['description'] = description
                    
                    # If Azure AI available, we could render and analyze the chart
                    # This would require chart rendering capability
                    
                    charts.append(chart_info)
        
        return charts
    
    def _describe_chart(self, chart) -> str:
        """Generate text description of a chart"""
        description_parts = []
        
        # Chart type
        chart_type = type(chart).__name__.replace('Chart', '')
        description_parts.append(f"This is a {chart_type} chart")
        
        # Title
        if hasattr(chart, 'title') and chart.title:
            description_parts.append(f"titled '{chart.title}'")
        
        # Series information
        if hasattr(chart, 'series'):
            series_count = len(list(chart.series))
            description_parts.append(f"with {series_count} data series")
        
        # Axes information
        if hasattr(chart, 'x_axis') and chart.x_axis and hasattr(chart.x_axis, 'title'):
            if chart.x_axis.title:
                description_parts.append(f"X-axis: {chart.x_axis.title}")
        
        if hasattr(chart, 'y_axis') and chart.y_axis and hasattr(chart.y_axis, 'title'):
            if chart.y_axis.title:
                description_parts.append(f"Y-axis: {chart.y_axis.title}")
        
        return ". ".join(description_parts) + "."
    
    def _extract_images(self, workbook) -> List[Dict[str, Any]]:
        """Extract images from workbook"""
        images = []
        
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_images'):
                for idx, img in enumerate(sheet._images):
                    image_info = {
                        'sheet': sheet.title,
                        'index': idx,
                        'anchor': str(img.anchor) if hasattr(img, 'anchor') else None
                    }
                    
                    try:
                        # Proper way to extract image data from openpyxl
                        if hasattr(img, 'ref'):
                            # img.ref is an Image object, get the actual bytes
                            if hasattr(img.ref, 'blob'):
                                image_data = img.ref.blob
                            elif hasattr(img.ref, 'data'):
                                image_data = img.ref.data()
                            else:
                                # Try to read from the image path
                                continue
                        elif hasattr(img, 'data'):
                            image_data = img.data()
                        else:
                            continue
                        
                        # Ensure we have bytes
                        if not isinstance(image_data, bytes):
                            continue
                        
                        # Process with Azure AI if available
                        if self.azure_client and isinstance(image_data, bytes):
                            # OCR the image
                            ocr_result = self.azure_client.process_image(image_data)
                            if ocr_result['success']:
                                image_info['extracted_text'] = ocr_result['text']
                            
                            # Analyze image content
                            analysis = self.azure_client.analyze_image(image_data)
                            if analysis['success']:
                                image_info['analysis'] = analysis['analysis']
                        
                        # Get image metadata
                        if PIL_AVAILABLE and isinstance(image_data, bytes):
                            img_obj = Image.open(BytesIO(image_data))
                            image_info['format'] = img_obj.format
                            image_info['size'] = img_obj.size
                            image_info['mode'] = img_obj.mode
                        
                        images.append(image_info)
                        
                    except Exception as e:
                        self.logger.error(f"Error processing image in sheet {sheet.title}: {e}")
        
        return images
    
    def _extract_hierarchical_data(self, workbook) -> Dict[str, Any]:
        """Extract hierarchical/floor manager data from workbook"""
        hierarchical_data = {}
        
        for sheet in workbook.worksheets:
            # Look for sheets that might contain org/floor data
            sheet_name_lower = sheet.title.lower()
            
            is_relevant = any(pattern in sheet_name_lower for pattern in 
                            self.floor_patterns + self.manager_patterns + ['org', 'hierarchy', 'structure'])
            
            if is_relevant:
                # Analyze sheet structure
                sheet_analysis = self._analyze_hierarchical_sheet(sheet)
                if sheet_analysis:
                    hierarchical_data[sheet.title] = sheet_analysis
        
        return hierarchical_data
    
    def _analyze_hierarchical_sheet(self, sheet) -> Optional[Dict[str, Any]]:
        """Analyze a sheet for hierarchical/organizational data"""
        if not PANDAS_AVAILABLE:
            return None
        
        try:
            # Convert sheet to DataFrame
            data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    data.append(list(row))
            
            if not data:
                return None
            
            df = pd.DataFrame(data)
            
            # Try to identify headers
            potential_header_row = 0
            for idx, row in df.iterrows():
                if row.notna().sum() > len(df.columns) * 0.5:  # More than 50% non-null
                    potential_header_row = idx
                    break
            
            # Set headers if found
            if potential_header_row > 0:
                df.columns = df.iloc[potential_header_row]
                df = df.iloc[potential_header_row + 1:].reset_index(drop=True)
            
            # Look for floor/manager related columns
            floor_cols = []
            manager_cols = []
            
            for col in df.columns:
                if col and isinstance(col, str):
                    col_lower = str(col).lower()
                    if any(pattern in col_lower for pattern in self.floor_patterns):
                        floor_cols.append(col)
                    if any(pattern in col_lower for pattern in self.manager_patterns):
                        manager_cols.append(col)
            
            if not (floor_cols or manager_cols):
                return None
            
            # Extract hierarchical structure
            hierarchy = {
                'type': 'floor_management',
                'structure': []
            }
            
            # Group by floor if floor columns exist
            if floor_cols:
                floor_col = floor_cols[0]
                grouped = df.groupby(floor_col)
                
                for floor, group in grouped:
                    floor_info = {
                        'floor': str(floor),
                        'employee_count': int(len(group))
                    }
                    
                    # Add manager info if available
                    if manager_cols:
                        managers = [str(m) for m in group[manager_cols[0]].dropna().unique().tolist()]
                        floor_info['managers'] = managers
                    
                    # Add department info if available
                    dept_cols = [col for col in df.columns if col and 'department' in str(col).lower()]
                    if dept_cols:
                        departments = [str(d) for d in group[dept_cols[0]].dropna().unique().tolist()]
                        floor_info['departments'] = departments
                    
                    hierarchy['structure'].append(floor_info)
            
            return hierarchy
            
        except Exception as e:
            self.logger.error(f"Error analyzing hierarchical sheet: {e}")
            return None
    
    def _extract_sheet_relationships(self, workbook) -> Dict[str, List[str]]:
        """Extract relationships between sheets based on formulas"""
        relationships = {}
        
        for sheet in workbook.worksheets:
            sheet_refs = set()
            
            # Scan formulas for sheet references
            for row in sheet.iter_rows():
                for cell in row:
                    if hasattr(cell, 'formula') and cell.formula:
                        # Look for sheet references in formula (e.g., Sheet2!A1)
                        for other_sheet in workbook.sheetnames:
                            if other_sheet != sheet.title and other_sheet in str(cell.formula):
                                sheet_refs.add(other_sheet)
            
            if sheet_refs:
                relationships[sheet.title] = list(sheet_refs)
        
        return relationships
    
    def _create_chunks(self, processed_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Create text chunks for vector storage with flat metadata structure"""
        chunks = []
        
        # Chunk sheet data
        for sheet in processed_data['sheets']:
            sheet_text = f"Excel Sheet: {sheet['name']}\n"
            sheet_text += f"Dimensions: {sheet['dimensions']}\n"
            
            # Add summary if available
            if 'summary' in sheet:
                sheet_text += f"Summary: {json.dumps(sheet['summary'])}\n"
            
            # Add sample data (first 10 rows)
            if sheet['data']:
                sheet_text += "Sample Data:\n"
                for row in sheet['data'][:10]:
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    sheet_text += f"{row_str}\n"
            
            # Add formulas if any
            if sheet['formulas']:
                sheet_text += f"\nFormulas found: {len(sheet['formulas'])}\n"
                for formula in sheet['formulas'][:5]:  # Sample formulas
                    sheet_text += f"  {formula['cell']}: {formula['formula']}\n"
            
            chunks.append({
                'text': sheet_text,
                'metadata': {
                    'source_type': 'excel_sheet',
                    'sheet_name': sheet['name'],
                    'content_type': 'sheet_data',
                    'dimensions': sheet['dimensions'],
                    'formula_count': len(sheet.get('formulas', [])),
                    'data_rows': len(sheet.get('data', []))
                }
            })
        
        # Chunk embedded object texts
        for obj in processed_data['embedded_objects']:
            if 'extracted_text' in obj and obj['extracted_text']:
                chunks.append({
                    'text': f"Embedded {obj['type']}: {obj['extracted_text']}",
                    'metadata': {
                        'source_type': 'embedded_object',
                        'object_type': obj['type'],
                        'content_type': 'embedded_content',
                        'object_name': obj.get('name', 'unknown'),
                        'object_size': obj.get('size', 0)
                    }
                })
        
        # Chunk image descriptions
        for img in processed_data['images']:
            img_text = f"Image in sheet {img['sheet']}"
            
            if 'extracted_text' in img:
                img_text += f"\nText found in image: {img['extracted_text']}"
            
            if 'analysis' in img and 'description' in img['analysis']:
                captions = img['analysis']['description'].get('captions', [])
                if captions:
                    img_text += f"\nImage description: {captions[0]['text']}"
            
            chunks.append({
                'text': img_text,
                'metadata': {
                    'source_type': 'excel_image',
                    'sheet_name': img['sheet'],
                    'content_type': 'image_content',
                    'image_index': img.get('index', 0),
                    'has_extracted_text': 'extracted_text' in img,
                    'has_analysis': 'analysis' in img
                }
            })
        
        # Chunk hierarchical data
        for sheet_name, hierarchy in processed_data['hierarchical_data'].items():
            if hierarchy and 'structure' in hierarchy:
                hier_text = f"Organizational Structure from {sheet_name}:\n"
                for floor_info in hierarchy['structure']:
                    hier_text += f"\nFloor {floor_info['floor']}:\n"
                    hier_text += f"  Employees: {floor_info['employee_count']}\n"
                    if 'managers' in floor_info:
                        hier_text += f"  Managers: {', '.join(floor_info['managers'])}\n"
                    if 'departments' in floor_info:
                        hier_text += f"  Departments: {', '.join(floor_info['departments'])}\n"
                
                chunks.append({
                    'text': hier_text,
                    'metadata': {
                        'source_type': 'excel_hierarchy',
                        'sheet_name': sheet_name,
                        'content_type': 'organizational_data',
                        'hierarchy_type': hierarchy.get('type', 'unknown'),
                        'structure_count': len(hierarchy.get('structure', []))
                    }
                })
        
        # Add file-level metadata chunk
        meta_text = f"Excel File: {processed_data['file_name']}\n"
        meta_text += f"Total Sheets: {len(processed_data['sheets'])}\n"
        meta_text += f"Embedded Objects: {len(processed_data['embedded_objects'])}\n"
        meta_text += f"Images: {len(processed_data['images'])}\n"
        
        if processed_data['metadata'].get('properties'):
            props = processed_data['metadata']['properties']
            if props.get('title'):
                meta_text += f"Title: {props['title']}\n"
            if props.get('creator'):
                meta_text += f"Creator: {props['creator']}\n"
            if props.get('description'):
                meta_text += f"Description: {props['description']}\n"
        
        chunks.append({
            'text': meta_text,
            'metadata': {
                'source_type': 'excel_metadata',
                'content_type': 'file_metadata',
                'total_sheets': len(processed_data['sheets']),
                'total_embedded_objects': len(processed_data['embedded_objects']),
                'total_images': len(processed_data['images']),
                'has_properties': bool(processed_data['metadata'].get('properties'))
            }
        })
        
        return chunks


# Convenience function for creating processor
def create_excel_processor(config: Optional[Dict[str, Any]] = None,
                         azure_config: Optional[Dict[str, Any]] = None) -> ExcelProcessor:
    """Create an Excel processor with optional Azure AI integration"""
    # Initialize Azure client if config provided
    azure_client = None
    if azure_config:
        try:
            from ..integrations.azure_ai.azure_client import AzureAIClient
            azure_client = AzureAIClient(azure_config)
            if not azure_client.is_available():
                logging.warning("Azure AI client created but not fully available")
        except Exception as e:
            logging.error(f"Failed to create Azure AI client: {e}")
    
    return ExcelProcessor(config=config, azure_client=azure_client)