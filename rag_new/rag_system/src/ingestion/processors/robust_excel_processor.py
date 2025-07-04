"""
Robust Excel Processor
Excel processor with comprehensive Azure AI integration and fallback
"""
import json
import logging
import os
from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
import io
import base64
from datetime import datetime

from .base_processor import BaseProcessor
from ...integrations.azure_ai.robust_azure_client import RobustAzureAIClient
from ...integrations.azure_ai.config_validator import AzureAIConfigValidator

try:
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    Image = None

class RobustExcelProcessor(BaseProcessor):
    """Excel processor with robust Azure AI integration"""
    
    def __init__(self, config: Dict[str, Any], azure_client: Optional[RobustAzureAIClient] = None):
        super().__init__(config)
        self.azure_client = azure_client
        self.logger = logging.getLogger('RobustExcelProcessor')
        
        # Debug configuration
        self.extraction_debug = os.getenv('RAG_EXTRACTION_DEBUG', 'false').lower() == 'true'
        self.save_extraction_dumps = os.getenv('RAG_SAVE_EXTRACTION_DUMPS', 'false').lower() == 'true'
        
        if self.extraction_debug:
            self.logger.setLevel(logging.DEBUG)
            self.logger.info("🔍 Robust Excel Processor - Debug mode enabled")
        
        # Configuration
        self.max_file_size_mb = config.get('max_file_size_mb', 50) or 50  # Ensure it's not None
        self.process_images = config.get('process_images', True)
        self.process_charts = config.get('process_charts', False)
        self.include_formulas = config.get('include_formulas', False)
        self.max_rows_per_sheet = config.get('max_rows_per_sheet', 10000) or 10000  # Ensure it's not None
        self.max_cols_per_sheet = config.get('max_cols_per_sheet', 100) or 100  # Ensure it's not None
        
        # Initialize Azure AI if not provided
        if not self.azure_client and config.get('azure_ai'):
            self._initialize_azure_client(config['azure_ai'])
        
        # Validate dependencies
        self._validate_dependencies()
    
    def _initialize_azure_client(self, azure_config: Dict[str, Any]):
        """Initialize Azure AI client with validation"""
        try:
            # Validate and fix configuration
            fixed_config, issues = AzureAIConfigValidator.validate_and_fix(azure_config)
            
            if issues:
                self.logger.info(f"Azure AI configuration fixes applied: {len(issues)} issues")
            
            # Check if we have enough configuration
            status = AzureAIConfigValidator.get_configuration_status(fixed_config)
            if not status['overall']['configuration_complete']:
                self.logger.warning("Azure AI configuration incomplete, will use fallback processing")
                return
            
            # Initialize client
            self.azure_client = RobustAzureAIClient(fixed_config)
            
            # Check if client is healthy
            if not self.azure_client.is_healthy():
                self.logger.warning("Azure AI services not healthy, will use fallback processing")
                service_status = self.azure_client.get_service_status()
                for service, details in service_status['services'].items():
                    if details['error']:
                        self.logger.warning(f"Azure {service} error: {details['error']}")
            else:
                available_services = self.azure_client.get_available_services()
                self.logger.info(f"Azure AI initialized successfully with services: {available_services}")
                
        except Exception as e:
            self.logger.error(f"Failed to initialize Azure AI: {e}")
            self.azure_client = None
    
    def _validate_dependencies(self):
        """Validate required dependencies"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing. Install with: pip install openpyxl")
        
        if self.process_images and not PIL_AVAILABLE:
            self.logger.warning("PIL not available, image processing will be limited")
    
    def can_process(self, file_path: str) -> bool:
        """Check if this processor can handle the file"""
        if not OPENPYXL_AVAILABLE:
            return False
        
        file_path = Path(file_path)
        return file_path.suffix.lower() in ['.xlsx', '.xlsm', '.xltx', '.xltm']
    
    def process(self, file_path: str, metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Process Excel file with comprehensive error handling"""
        start_time = datetime.now()
        file_path = Path(file_path)
        
        self.logger.info(f"📊 Starting Excel processing: {file_path}")
        self.logger.info(f"📊 File type: EXCEL | File size: {file_path.stat().st_size:,} bytes")
        self.logger.info(f"📊 Processor: RobustExcelProcessor")
        
        # Validate file
        validation_result = self._validate_file(file_path)
        if not validation_result['valid']:
            self.logger.error(f"📊 Excel validation failed: {validation_result['error']}")
            return {
                'success': False,
                'error': validation_result['error'],
                'processor': 'robust_excel'
            }
        
        self.logger.info(f"📊 Excel file validation passed")
        
        try:
            # Load workbook
            self.logger.debug(f"📊 Loading Excel workbook...")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # ADD LOGGING: Workbook properties
            self.logger.info(f"📊 Excel Workbook Properties:")
            self.logger.info(f"  - Worksheets: {len(workbook.sheetnames)}")
            self.logger.info(f"  - Sheet names: {workbook.sheetnames}")
            self.logger.info(f"  - Azure AI available: {'Yes' if (self.azure_client and self.azure_client.is_healthy()) else 'No'}")
            
            # Process workbook
            result = self._process_workbook(workbook, file_path, metadata)
            # Close workbook to release file handle (important on Windows)
            try:
                workbook.close()
            except Exception as _close_err:
                self.logger.debug(f"Workbook close failed: {_close_err}")
            
            # Processing summary
            processing_time = (datetime.now() - start_time).total_seconds()
            
            # Add processing metadata
            result.update({
                'success': True,
                'processor': 'robust_excel',
                'file_size': file_path.stat().st_size,
                'processing_start': start_time.isoformat(),
                'processing_end': datetime.now().isoformat(),
                'processing_time_seconds': processing_time,
                'azure_ai_used': self.azure_client is not None and self.azure_client.is_healthy()
            })
            
            self.logger.info(f"✅ Excel processing completed in {processing_time:.2f}s")
            self.logger.info(f"📊 Final Result:")
            self.logger.info(f"  - Total text length: {len(result.get('text', '')):,} chars")
            self.logger.info(f"  - Sheets processed: {len(result.get('sheets', []))}")
            self.logger.info(f"  - Images processed: {result.get('images_processed', 0)}")
            self.logger.info(f"  - Charts processed: {result.get('charts_processed', 0)}")
            self.logger.info(f"  - Total cells: {result.get('total_cells', 0):,}")
            
            # ADD LOGGING: Save debug data if enabled
            if self.save_extraction_dumps or self.logger.isEnabledFor(logging.DEBUG):
                debug_file = f"debug_excel_extract_{file_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                debug_path = Path("data/logs") / debug_file
                debug_path.parent.mkdir(parents=True, exist_ok=True)
                
                # Create a serializable version
                debug_result = {k: v for k, v in result.items()}
                debug_result['chunk_preview'] = [
                    {
                        'sheet_name': sheet.get('name', 'Unknown'),
                        'text_length': len(sheet.get('text', '')),
                        'text_preview': sheet.get('text', '')[:200] + '...' if len(sheet.get('text', '')) > 200 else sheet.get('text', ''),
                        'cell_count': sheet.get('cell_count', 0)
                    }
                    for sheet in result.get('sheets', [])[:3]  # First 3 sheets
                ]
                
                try:
                    with open(debug_path, 'w', encoding='utf-8') as f:
                        json.dump(debug_result, f, indent=2, ensure_ascii=False, default=str)
                    self.logger.debug(f"Full Excel extraction saved to {debug_path}")
                except Exception as e:
                    self.logger.warning(f"Failed to save debug data: {e}")
            
            return result
            
        except Exception as e:
            error_msg = f"Excel processing failed: {str(e)}"
            processing_time = (datetime.now() - start_time).total_seconds()
            self.logger.error(f"❌ {error_msg} (after {processing_time:.2f}s)")
            return {
                'success': False,
                'error': error_msg,
                'processor': 'robust_excel',
                'processing_time_seconds': processing_time
            }
    
    def _validate_file(self, file_path: Path) -> Dict[str, Any]:
        """Validate Excel file"""
        if not file_path.exists():
            return {'valid': False, 'error': f"File not found: {file_path}"}
        
        # Check file size - ensure max_file_size_mb is not None
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        max_size_mb = self.max_file_size_mb or 50  # Fallback to 50MB if None
        
        if file_size_mb > max_size_mb:
            return {
                'valid': False, 
                'error': f"File too large: {file_size_mb:.1f}MB > {max_size_mb}MB"
            }
        
        # Check file extension
        if not self.can_process(str(file_path)):
            return {'valid': False, 'error': f"Unsupported file type: {file_path.suffix}"}
        
        # Try to open file
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            workbook.close()
        except Exception as e:
            return {'valid': False, 'error': f"Cannot open Excel file: {str(e)}"}
        
        return {'valid': True}
    
    def _process_workbook(self, workbook, file_path: Path, metadata: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """Process entire workbook"""
        self.logger.info(f"📊 Processing Excel workbook with {len(workbook.sheetnames)} sheets")
        
        result = {
            'text': '',
            'sheets': [],
            'images_processed': 0,
            'charts_processed': 0,
            'total_cells': 0,
            'processing_stats': {}
        }
        
        # Process each worksheet
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
            sheet_start = datetime.now()
            
            try:
                self.logger.debug(f"📊 Processing sheet {sheet_idx}/{len(workbook.sheetnames)}: '{sheet_name}'")
                sheet = workbook[sheet_name]
                sheet_result = self._process_worksheet(sheet, sheet_name)
                
                # ADD LOGGING: Sheet processing results
                sheet_processing_time = (datetime.now() - sheet_start).total_seconds()
                self.logger.info(f"📊 Sheet '{sheet_name}' processed:")
                self.logger.info(f"  - Text length: {len(sheet_result['text']):,} chars")
                self.logger.info(f"  - Cell count: {sheet_result.get('cell_count', 0):,}")
                self.logger.info(f"  - Images: {sheet_result.get('images_processed', 0)}")
                self.logger.info(f"  - Charts: {sheet_result.get('charts_processed', 0)}")
                self.logger.info(f"  - Processing time: {sheet_processing_time:.2f}s")
                
                # ADD DETAILED CONTENT LOGGING like PDF processor
                if self.logger.isEnabledFor(logging.DEBUG):
                    self.logger.debug(f"  - Dimensions: {sheet_result.get('metadata', {}).get('dimensions', 'Unknown')}")
                    self.logger.debug(f"  - Has merged cells: {sheet_result.get('metadata', {}).get('has_merged_cells', False)}")
                    self.logger.debug(f"  - Text preview: {sheet_result['text'][:200]}..." if len(sheet_result['text']) > 200 else f"  - Text: {sheet_result['text']}")
                    
                    # Log sample cell values for debugging
                    if 'sample_cells' in sheet_result.get('metadata', {}):
                        sample_cells = sheet_result['metadata']['sample_cells']
                        self.logger.debug(f"  - Sample cells: {sample_cells}")
                    
                    # Log formulas if any
                    if 'formula_count' in sheet_result.get('metadata', {}):
                        formula_count = sheet_result['metadata']['formula_count']
                        self.logger.debug(f"  - Formulas found: {formula_count}")
                        
                    # Log data types detected
                    if 'data_types' in sheet_result.get('metadata', {}):
                        data_types = sheet_result['metadata']['data_types']
                        self.logger.debug(f"  - Data types: {data_types}")
                
                # ADD LOGGING: For table-like structures
                if 'tables' in sheet_result and sheet_result['tables']:
                    for table_idx, table in enumerate(sheet_result['tables']):
                        self.logger.info(f"  - Table {table_idx + 1}: {table.get('rows', 0)} rows, {table.get('cols', 0)} cols")
                        if self.logger.isEnabledFor(logging.DEBUG):
                            self.logger.debug(f"    Table preview: {table.get('text', '')[:100]}...")
                
                # ADD LOGGING: For image OCR results
                if 'images' in sheet_result and sheet_result['images']:
                    for img_idx, img in enumerate(sheet_result['images']):
                        if img.get('ocr_text'):
                            self.logger.info(f"  - Image {img_idx + 1} OCR: {len(img['ocr_text'])} chars extracted")
                            if self.logger.isEnabledFor(logging.DEBUG):
                                self.logger.debug(f"    OCR Text Preview: {img['ocr_text'][:200]}...")
                        else:
                            self.logger.debug(f"  - Image {img_idx + 1}: No OCR text extracted")

                result['sheets'].append(sheet_result)
                result['text'] += f"\n\n--- Sheet: {sheet_name} ---\n"
                result['text'] += sheet_result['text']
                result['images_processed'] += sheet_result.get('images_processed', 0)
                result['charts_processed'] += sheet_result.get('charts_processed', 0)
                result['total_cells'] += sheet_result.get('cell_count', 0)
                
            except Exception as e:
                sheet_processing_time = (datetime.now() - sheet_start).total_seconds()
                self.logger.error(f"❌ Failed to process sheet '{sheet_name}' (after {sheet_processing_time:.2f}s): {e}")
                result['sheets'].append({
                    'name': sheet_name,
                    'error': str(e),
                    'text': f"[Error processing sheet {sheet_name}: {str(e)}]",
                    'processing_time_seconds': sheet_processing_time
                })
        
        # Process workbook-level elements
        if hasattr(workbook, '_images') and self.process_images:
            workbook_images = self._process_workbook_images(workbook)
            result['images_processed'] += workbook_images['count']
            if workbook_images['text']:
                result['text'] += f"\n\n--- Workbook Images ---\n{workbook_images['text']}"
        
        # Add summary statistics
        result['processing_stats'] = {
            'total_sheets': len(workbook.sheetnames),
            'processed_sheets': len([s for s in result['sheets'] if 'error' not in s]),
            'failed_sheets': len([s for s in result['sheets'] if 'error' in s]),
            'azure_ai_available': self.azure_client is not None and self.azure_client.is_healthy()
        }
        
        return result
    
    def _process_worksheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Process individual worksheet"""
        result = {
            'name': sheet_name,
            'text': '',
            'cell_count': 0,
            'images_processed': 0,
            'charts_processed': 0,
            'tables': [],
            'metadata': {}
        }
        
        # Get sheet dimensions
        max_row = min(sheet.max_row, self.max_rows_per_sheet)
        max_col = min(sheet.max_column, self.max_cols_per_sheet)
        
        self.logger.debug(f"📊 Sheet '{sheet_name}' dimensions: {max_row}x{max_col} (original: {sheet.max_row}x{sheet.max_column})")
        
        if max_row == 0 or max_col == 0:
            self.logger.debug(f"📊 Sheet '{sheet_name}' is empty")
            result['text'] = f"[Empty sheet: {sheet_name}]"
            return result
        
        # Extract cell data
        cell_data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            row_data = []
            for cell in row:
                cell_value = self._extract_cell_value(cell)
                if cell_value:
                    row_data.append(str(cell_value))
                    result['cell_count'] += 1
                else:
                    row_data.append('')
            
            # Only add non-empty rows
            if any(row_data):
                cell_data.append(row_data)
        
        # Convert to text
        if cell_data:
            # Try to detect if first row is headers
            if self._looks_like_headers(cell_data[0]):
                result['text'] = self._format_as_table(cell_data)
            else:
                result['text'] = self._format_as_text(cell_data)
        
        # Process images
        if self.process_images and hasattr(sheet, '_images'):
            images_result = self._process_sheet_images(sheet, sheet_name)
            result['images_processed'] = images_result['count']
            if images_result['text']:
                result['text'] += f"\n\n--- Images in {sheet_name} ---\n{images_result['text']}"
        
        # Process charts
        if self.process_charts and hasattr(sheet, '_charts'):
            charts_result = self._process_sheet_charts(sheet, sheet_name)
            result['charts_processed'] = charts_result['count']
            if charts_result['text']:
                result['text'] += f"\n\n--- Charts in {sheet_name} ---\n{charts_result['text']}"
        
        # Add sheet metadata
        result['metadata'] = {
            'dimensions': f"{max_row}x{max_col}",
            'has_merged_cells': len(sheet.merged_cells.ranges) > 0,
            'has_images': hasattr(sheet, '_images') and len(sheet._images) > 0,
            'has_charts': hasattr(sheet, '_charts') and len(sheet._charts) > 0
        }
        
        return result
    
    def _extract_cell_value(self, cell) -> Any:
        """Extract value from cell with formula handling"""
        if cell.value is None:
            return None
        
        # Handle formulas
        if self.include_formulas and hasattr(cell, 'formula') and cell.formula:
            return f"{cell.value} (Formula: {cell.formula})"
        
        return cell.value
    
    def _looks_like_headers(self, row: List[str]) -> bool:
        """Detect if row looks like table headers"""
        if not row:
            return False
        
        # Check if most cells are non-empty strings
        non_empty = [cell for cell in row if cell and isinstance(cell, str)]
        if len(non_empty) < len(row) * 0.7:
            return False
        
        # Check if they look like headers (short, descriptive)
        avg_length = sum(len(str(cell)) for cell in non_empty) / len(non_empty)
        return avg_length < 30  # Headers are usually short
    
    def _format_as_table(self, data: List[List[str]]) -> str:
        """Format data as a structured table"""
        if not data:
            return ""
        
        # Use first row as headers
        headers = data[0]
        rows = data[1:]
        
        table_text = f"Table with columns: {', '.join(headers)}\n\n"
        
        for i, row in enumerate(rows[:100]):  # Limit to first 100 rows
            row_items = []
            for j, (header, value) in enumerate(zip(headers, row)):
                if value:
                    row_items.append(f"{header}: {value}")
            
            if row_items:
                table_text += f"Row {i+1}: {'; '.join(row_items)}\n"
        
        if len(rows) > 100:
            table_text += f"\n... and {len(rows) - 100} more rows"
        
        return table_text
    
    def _format_as_text(self, data: List[List[str]]) -> str:
        """Format data as plain text"""
        text_lines = []
        for row in data:
            non_empty_cells = [cell for cell in row if cell]
            if non_empty_cells:
                text_lines.append(' | '.join(non_empty_cells))
        
        return '\n'.join(text_lines)
    
    def _process_sheet_images(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Process images in worksheet"""
        result = {'count': 0, 'text': ''}
        
        if not hasattr(sheet, '_images') or not sheet._images:
            return result
        
        for image in sheet._images:
            try:
                image_result = self._process_embedded_image(image, f"{sheet_name}_image_{result['count']}")
                if image_result['success']:
                    result['text'] += f"\nImage {result['count'] + 1}: {image_result['text']}\n"
                else:
                    result['text'] += f"\nImage {result['count'] + 1}: [Processing failed - {image_result.get('error', 'Unknown error')}]\n"
                
                result['count'] += 1
                
            except Exception as e:
                self.logger.error(f"Failed to process image in {sheet_name}: {e}")
                result['text'] += f"\nImage {result['count'] + 1}: [Error: {str(e)}]\n"
                result['count'] += 1
        
        return result
    
    def _process_workbook_images(self, workbook) -> Dict[str, Any]:
        """Process workbook-level images"""
        result = {'count': 0, 'text': ''}
        
        # This would need to be implemented based on openpyxl's image handling
        # For now, return empty result
        return result
    
    def _process_sheet_charts(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Process charts in worksheet"""
        result = {'count': 0, 'text': ''}
        
        if not hasattr(sheet, '_charts') or not sheet._charts:
            return result
        
        for chart in sheet._charts:
            try:
                chart_text = self._extract_chart_data(chart)
                result['text'] += f"\nChart {result['count'] + 1}: {chart_text}\n"
                result['count'] += 1
                
            except Exception as e:
                self.logger.error(f"Failed to process chart in {sheet_name}: {e}")
                result['text'] += f"\nChart {result['count'] + 1}: [Error: {str(e)}]\n"
                result['count'] += 1
        
        return result
    
    def _extract_chart_data(self, chart) -> str:
        """Extract data from chart object"""
        try:
            chart_info = []
            
            # Basic chart information
            if hasattr(chart, 'title') and chart.title:
                chart_info.append(f"Title: {chart.title}")
            
            if hasattr(chart, 'type'):
                chart_info.append(f"Type: {chart.type}")
            
            # Try to extract data series information
            if hasattr(chart, 'series'):
                series_info = []
                for series in chart.series:
                    if hasattr(series, 'title') and series.title:
                        series_info.append(f"Series: {series.title}")
                
                if series_info:
                    chart_info.append(f"Data series: {', '.join(series_info)}")
            
            return '; '.join(chart_info) if chart_info else "Chart (no details available)"
            
        except Exception as e:
            return f"Chart (extraction failed: {str(e)})"
    
    def _process_embedded_image(self, image, image_name: str) -> Dict[str, Any]:
        """Process embedded image with Azure AI or fallback"""
        try:
            # Extract image data
            image_data = self._extract_image_data(image)
            if not image_data:
                return {
                    'success': False,
                    'error': 'Could not extract image data',
                    'text': '[Image data unavailable]'
                }
            
            # Process with Azure AI if available
            if self.azure_client and self.azure_client.is_healthy():
                return self.azure_client.process_image_with_fallback(
                    image_data,
                    fallback_handler=self._local_image_processing
                )
            else:
                # Use local processing
                return self._local_image_processing(image_data)
                
        except Exception as e:
            self.logger.error(f"Failed to process embedded image {image_name}: {e}")
            return {
                'success': False,
                'error': str(e),
                'text': f'[Image processing failed: {str(e)}]'
            }
    
    def _extract_image_data(self, image) -> Optional[bytes]:
        """Extract binary data from openpyxl image object"""
        try:
            # Proper way to extract image data from openpyxl
            if hasattr(image, 'ref'):
                # img.ref is an Image object, get the actual bytes
                if hasattr(image.ref, 'blob'):
                    image_data = image.ref.blob
                elif hasattr(image.ref, 'data'):
                    image_data = image.ref.data()
                else:
                    # Try to read from the image path
                    return None
            elif hasattr(image, '_data'):
                image_data = image._data()
            else:
                self.logger.warning("Could not extract image data - unknown format")
                return None
            
            # Ensure we have bytes
            if not isinstance(image_data, bytes):
                return None
                
            return image_data
            
        except Exception as e:
            self.logger.error(f"Failed to extract image data: {e}")
            return None
    
    def _local_image_processing(self, image_data: bytes) -> Dict[str, Any]:
        """Local image processing fallback"""
        try:
            if not PIL_AVAILABLE:
                return {
                    'success': True,
                    'text': '[Embedded image - processing not available]',
                    'service': 'none'
                }
            
            # Basic image metadata extraction
            image = Image.open(io.BytesIO(image_data))
            
            # Try OCR if available
            try:
                import pytesseract
                text = pytesseract.image_to_string(image)
                
                return {
                    'success': True,
                    'text': text if text.strip() else '[Image with no readable text]',
                    'caption': f'Image ({image.size[0]}x{image.size[1]} pixels)',
                    'service': 'pytesseract',
                    'metadata': {
                        'width': image.size[0],
                        'height': image.size[1],
                        'mode': image.mode,
                        'format': image.format
                    }
                }
            except ImportError:
                # No OCR available
                return {
                    'success': True,
                    'text': f'[Image: {image.size[0]}x{image.size[1]} pixels, {image.mode} mode]',
                    'caption': 'Embedded image',
                    'service': 'pil_metadata',
                    'metadata': {
                        'width': image.size[0],
                        'height': image.size[1],
                        'mode': image.mode,
                        'format': image.format
                    }
                }
            except Exception as e:
                self.logger.error(f"OCR processing failed: {e}")
                return {
                    'success': True,
                    'text': f'[Image: {image.size[0]}x{image.size[1]} pixels - text extraction failed]',
                    'service': 'pil_basic'
                }
                
        except Exception as e:
            self.logger.error(f"Local image processing failed: {e}")
            return {
                'success': False,
                'text': '[Image processing failed]',
                'error': str(e),
                'service': 'none'
            }
    
    def get_supported_extensions(self) -> List[str]:
        """Get list of supported file extensions"""
        return ['.xlsx', '.xlsm', '.xltx', '.xltm']
    
    def get_processor_info(self) -> Dict[str, Any]:
        """Get processor information and capabilities"""
        return {
            'name': 'robust_excel',
            'description': 'Robust Excel processor with Azure AI integration',
            'supported_extensions': self.get_supported_extensions(),
            'capabilities': {
                'text_extraction': True,
                'image_processing': self.process_images,
                'chart_processing': self.process_charts,
                'formula_extraction': self.include_formulas,
                'azure_ai_integration': self.azure_client is not None,
                'azure_ai_healthy': self.azure_client.is_healthy() if self.azure_client else False
            },
            'dependencies': {
                'openpyxl': OPENPYXL_AVAILABLE,
                'pil': PIL_AVAILABLE,
                'azure_ai': self.azure_client is not None
            },
            'configuration': {
                'max_file_size_mb': self.max_file_size_mb,
                'max_rows_per_sheet': self.max_rows_per_sheet,
                'max_cols_per_sheet': self.max_cols_per_sheet,
                'process_images': self.process_images,
                'process_charts': self.process_charts,
                'include_formulas': self.include_formulas
            }
        } 